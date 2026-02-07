import pandas as pd
import openpyxl
import xlrd
import os

file_path = r'F:\OCB-8.xls'

print(f"File size: {os.path.getsize(file_path)} bytes")

def try_read_excel(path):
    # Try 1: xlrd
    try:
        print("\nTrying xlrd...")
        book = xlrd.open_workbook(path)
        print(f"xlrd success! Sheets: {book.sheet_names()}")
        return
    except Exception as e:
        print(f"xlrd failed: {e}")

    # Try 2: openpyxl (assuming it might be xlsx in disguise)
    try:
        print("\nTrying openpyxl...")
        wb = openpyxl.load_workbook(path, data_only=True)
        print(f"openpyxl success! Sheets: {wb.sheetnames}")
        sheet = wb.active
        for row in sheet.iter_rows(max_row=50, max_col=15):
            r_data = []
            for cell in row:
                if cell.value is not None:
                    r_data.append(f"{cell.column_letter}{cell.row}: {cell.value}")
            if r_data:
                print(" | ".join(r_data))
        return
    except Exception as e:
        print(f"openpyxl failed: {e}")

    # Try 3: pandas with engine='xlrd'
    try:
        print("\nTrying pandas xlrd...")
        df = pd.read_excel(path, engine='xlrd')
        print("pandas xlrd success!")
        print(df.head())
        return
    except Exception as e:
        print(f"pandas xlrd failed: {e}")

try_read_excel(file_path)
