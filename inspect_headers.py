import openpyxl

file_path = 'DİZAYN CIKTISI.xlsx'
try:
    wb = openpyxl.load_workbook(file_path, data_only=True)
    sheet = wb.active
    print(f"Reading top rows of {file_path}...")
    
    for row in range(1, 30):
        line_data = []
        for col in range(1, 10): # A to I
            val = sheet.cell(row=row, column=col).value
            if val:
                line_data.append(f"{openpyxl.utils.get_column_letter(col)}{row}: {val}")
        if line_data:
            print(" | ".join(line_data))
            
except Exception as e:
    print(f"Error: {e}")
